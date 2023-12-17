# Import libraries
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import pandas as pd
import time
import datetime
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Path to local chromedriver for selenium
driver = webdriver.Chrome(executable_path='C:/Users/Tommy Chu/Documents/chromedriver.exe')
driver.set_page_load_timeout(10)

# Declare a function that passes through the URL that will be scraped
def getPrices(url):
    
    driver.get(url)
    rawText = driver.find_element(By.XPATH, '/html/body/pre').text
    # Drop the symbols
    symbols = '"[]:{},-'
    bareText = rawText
    for char in symbols:
        bareText = bareText.replace(char, " ")
    
    #Tokenize raw text
    tokens = bareText.split(' ')
    tokens = [ele for ele in tokens if ele != '']
    
    priceList = []
    dateList = []
    
    for idx, x in enumerate(tokens):
        if x == "last":
            price = tokens[idx + 1]
            for y in range(idx, len(tokens)):
                if tokens[y] == "expirationMonth":
                    month = int(datetime.datetime.strptime(tokens[y+1], "%b").month)
                    year = int(tokens[y+2])
                    date = datetime.datetime(year,month,1)
                    break
            priceList.append(price)
            dateList.append(date)
                
    results = pd.DataFrame([dateList, priceList]).T

    return results 

# Initialize list of links for the various commodity contracts
HH = getPrices('https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/444/G?quoteCodes=null&_=1560171518204')
#WTI = getPrices('https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/4707/G?quoteCodes=null&_=1560171518204')
WTI = getPrices('https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/425/G?quoteCodes=null&_=1560171518204')
WCS = getPrices('https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/8289/G?quoteCodes=null&_=1560171518204')
MSW = getPrices('https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/8290/G?quoteCodes=null&_=1560171518204')
Cond = getPrices('https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/8062/G?quoteCodes=null&_=1560171518204')
Brent = getPrices('https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/424/G?quoteCodes=null&_=1560171518204')
JKM = getPrices('https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/7049/G?quoteCodes=null&_=1560171518204')
FX = getPrices('https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/48/G?quoteCodes=null&_=1560171518204')

HH = HH.set_axis(['Contract Date', 'Price'], axis=1, inplace=False)
WTI = WTI.set_axis(['Contract Date', 'Price'], axis=1, inplace=False)
WCS = WCS.set_axis(['Contract Date', 'Price'], axis=1, inplace=False)
MSW = MSW.set_axis(['Contract Date', 'Price'], axis=1, inplace=False)
Cond = Cond.set_axis(['Contract Date', 'Price'], axis=1, inplace=False)
Brent = Brent.set_axis(['Contract Date', 'Price'], axis=1, inplace=False)
JKM = JKM.set_axis(['Contract Date', 'Price'], axis=1, inplace=False)
FX = FX.set_axis(['Contract Date', 'Price'], axis=1, inplace=False)

# Label columns for the date it was scraped
HH['Date'] = date.today()
WTI['Date'] = date.today()
WCS['Date'] = date.today()
MSW['Date'] = date.today()
Cond['Date'] = date.today()
Brent['Date'] = date.today()
JKM['Date'] = date.today()
FX['Date'] = date.today()

# Load output file to save scraped data
book = load_workbook(r'C:\Users\Tommy Chu\Dropbox\(3) Python\Commodity Price Scrapes\priceOutputCME.xlsx')
writer = pd.ExcelWriter(r'C:\Users\Tommy Chu\Dropbox\(3) Python\Commodity Price Scrapes\priceOutputCME.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

HH = HH.copy()
HH.to_excel(writer, sheet_name='HH', startrow = writer.sheets['HH'].max_row, header=False)

WTI = WTI.copy()
WTI.to_excel(writer, sheet_name='WTI', startrow = writer.sheets['WTI'].max_row, header=False)

WCS = WCS.copy()
WCS.to_excel(writer, sheet_name='WCS', startrow = writer.sheets['WCS'].max_row, header=False)

MSW = MSW.copy()
MSW.to_excel(writer, sheet_name='MSW', startrow = writer.sheets['MSW'].max_row, header=False)

Cond = Cond.copy()
Cond.to_excel(writer, sheet_name='Cond', startrow = writer.sheets['Cond'].max_row, header=False)

Brent = Brent.copy()
Brent.to_excel(writer, sheet_name='Brent', startrow = writer.sheets['Brent'].max_row, header=False)

JKM = JKM.copy()
JKM.to_excel(writer, sheet_name='JKM', startrow = writer.sheets['JKM'].max_row, header=False)

FX = FX.copy()
FX.to_excel(writer, sheet_name='FX', startrow = writer.sheets['FX'].max_row, header=False)

writer.save()
time.sleep(2)
driver.close()

