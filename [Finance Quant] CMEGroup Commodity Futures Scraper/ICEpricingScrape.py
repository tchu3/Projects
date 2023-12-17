# Import libraries
from selenium import webdriver
from selenium.webdriver.common.by import By
import requests
import pandas as pd
from datetime import date
from openpyxl import load_workbook
import time

url = "https://www.theice.com/marketdata/public-web/ngx/daily-settlement-price/report"

# Declare lists for the ICE contracts to be scraped
contracts = [
        ["AECO","NGX Fin BS, LD1 for 5A, (US/MM), AB-NIT"],
        ["Dawn","NGX Phys, BS, LD1 (US/MM), Union-Dawn"],
        ["Emerson 1","NGX Phys, BS, LD1 (US/MM), TCPL - Emerson 1"],
        ["Emerson 2","NGX Phys, BS, LD1 (US/MM), TCPL - Emerson 2"],
        ["Sumas","NGX Phys, BS, LD1 (US/MM), Spectra - Hunt"],
        ["Chicago CG","NG Firm Phys, BS, LD1, (US/MM), NGPL-Nicor"],
        ["Malin","NG Firm Phys, BS, LD1, (US/MM), GTN-Malin"],
        ["PGE Citygate","NG Firm Phys, BS, LD1, (US/MM), PGE-Citygate"],
        ["PGE Topock","NG Firm Phys, BS, LD1, (US/MM), PGE-Topock"],
        ["SoCal Citygate","NG Firm Phys, BS, LD1, (US/MM), SoCal-Citygate"],
        ["Waha","NG Firm Phys, BS, LD1, (US/MM), Waha"],
        ["Cheyenne","NG Firm Phys, BS, LD1, (US/MM), REX Cheyenne"],
        ["Michcon","NG Firm Phys, BS, LD1, (US/MM), Michcon"],
        ["NGPL-TXOK","NG Firm Phys, BS, LD1, (US/MM), NGPL-TXOK East"],
        ["Ventura","NG Firm Phys, BS, LD1, (US/MM), NNG-Ventura"],
        ["TETCO-M1","NG Firm Phys, BS, LD1, (US/MM), TETCO-M1 30"],
        ["TETCO-M2","NG Firm Phys, BS, LD1, (US/MM), TETCO-M2 (receipts)"],
        ["TETCO-M3","NG Firm Phys, BS, LD1, (US/MM), TETCO-M3"],
        ["TETCO-ELA","NG Firm Phys, BS, LD1, (US/MM), TETCO-ELA"],
        ["TETCO-STX","NG Firm Phys, BS, LD1, (US/MM), TETCO-STX"],
        ["TETCO-WLA","NG Firm Phys, BS, LD1, (US/MM), TETCO-WLA"],
        ["Transco-Z5","NG Firm Phys, BS, LD1, (US/MM), Transco-Z5 South"],
        ["Transco-Z6","NG Firm Phys, BS, LD1, (US/MM), Transco-Z6 non-NY"],
        ["Transco-Z4","NG Firm Phys, BS, LD1, (US/MM), Transco-85"]]
        
modContracts = [
        ["Empress","NGX Phys, FP (CA/GJ), TCPL - Empress"],
        ["Station 2","NGX Phys, FP (US/MM), Spectra - Stn 2"]]

# Iterate through the main contracts 
for x in contracts:
    print(x[0])
    # Data to pass through to the JSON
    data = {
        "reportId": "254",
        "group": "Natural Gas",
        "selectedMarket": x[1],
        "selectedTimePeriod": "Current Settlement",
        "rcMode": "2",
    }
    
    df = pd.read_html(requests.post(url, data=data).text)[0]
    df['Date'] = date.today()
    
    # Write to the output file
    book = load_workbook(r'C:\Users\Tommy Chu\Dropbox\(3) Python\Commodity Price Scrapes\priceOutputICE.xlsx')
    writer = pd.ExcelWriter(r'C:\Users\Tommy Chu\Dropbox\(3) Python\Commodity Price Scrapes\priceOutputICE.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    
    df = df[['BeginDate','Settle','Date']].copy()
    df.to_excel(writer, sheet_name=x[0], startrow = writer.sheets[x[0]].max_row, header=False)
    writer.save()

# Iterate through the contracts that require modifying
for x in modContracts:
    print(x[0])
    # Data to pass through to the JSON
    HHData = {
        "reportId": "254",
        "group": "Natural Gas",
        "selectedMarket": "NGX Fin FF, FP for LD1, (US/MM), Henry",
        "selectedTimePeriod": "Current Settlement",
        "rcMode": "2",
    }
    HH = pd.read_html(requests.post(url, data=HHData).text)[0]
    
    data = {
        "reportId": "254",
        "group": "Natural Gas",
        "selectedMarket": x[1],
        "selectedTimePeriod": "Current Settlement",
        "rcMode": "2",
    }
    
    df = pd.read_html(requests.post(url, data=data).text)[0]
    
    # Write to the output file
    book = load_workbook(r'C:\Users\Tommy Chu\Dropbox\(3) Python\Commodity Price Scrapes\priceOutputICE.xlsx')
    writer = pd.ExcelWriter(r'C:\Users\Tommy Chu\Dropbox\(3) Python\Commodity Price Scrapes\priceOutputICE.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    
    df = df.merge(HH, on='BeginDate', how='left')
    
    # Calculate basis using the HH future prices
    # Modify the Empress contract by turning it into USD/MMbtu
    if x[0] == "Empress":
        df['Settle'] = df['Settle_x']*0.73/0.94781712031332 - df['Settle_y'] 
    else:
        df['Settle'] = (df['Settle_x'] - df['Settle_y']) 
    
    # Take the average for the front months since they will be in days
    df['Month'] = pd.to_datetime(df['BeginDate']).dt.month
    df['Year'] = pd.to_datetime(df['BeginDate']).dt.year
    df = df.groupby(['Year','Month'],as_index=False).mean()
    df['dummy'] = str("01")
    df['BeginDate'] = df[['Year', 'Month', 'dummy']].astype(str).agg('-'.join, axis=1)
    df['Date'] = date.today()
    
    df = df.dropna()
    df = df[['BeginDate','Settle','Date']].copy()
    df.to_excel(writer, sheet_name=x[0], startrow = writer.sheets[x[0]].max_row, header=False)
    writer.save()
    
time.sleep(2)

    